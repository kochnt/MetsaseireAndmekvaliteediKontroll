import pandas as pd
import numpy as np
import openpyxl as op
import time
import sys
from openpyxl.utils import get_column_letter
from excel_df import excel_df
# Vajalikud moodulid
from exists_check import exists_check
from valid_names_check import valid_names_check 
from format_seirekoha_kkr_check import format_seirekoha_kkr_check
from correlation_check import correlation_check
from condition_seirekoha_x_check import condition_seirekoha_x_check
from condition_seirekoha_y_check import condition_seirekoha_y_check
from min_max_check import min_max_check
from condition_vaartus_check import condition_vaartus_check

#====================================================================================================================
# Algusaeg
start_time = time.time()
#----------------------------------------------------------------------------------------------------------

sheet_name = 'Metsaseire_2022'

print(f"\nTabel {sheet_name}")

# Tekstifail, kuhu kirjutan tulemused

output_file = f'results_{sheet_name}.txt' 
table = excel_df(sheet_name)
if isinstance(table, bool):
            print(f"Tabel on tüüpis: {type(table).__name__}, mitte pandas DataFrame")
            print(f"Andmete lugemisel tekkis viga")
            sys.exit()
else:
    total_count = len(table) # Ridade koguarv tabelis

#total_count = len(table) # Ridade koguarv tabelis

table_str = f"\nTabel {sheet_name} \nRidade koguarv: {total_count}\n"
with open(output_file, 'a', encoding='utf-8') as file:
    file.write(table_str) # Kirjutan muutujast info faili

#====================================================================================================================

# Excel fail, kuhu loon tulemuste tabeli
excel_file = 'Andmekvaliteet.xlsx'
sheet_name = 'Metsaseire_2022_Andmekvaliteet' # lehe nimi

try:
    wb = op.load_workbook(excel_file)
    

    # Õige lehe valimine pealkirja järgi
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name) # Loon uus leht, kui seda pole
        sheet = wb[sheet_name]
        headers = ["Reegli ärivõti", "Dimensioon", "Probleemi liik", "Reegli kirjeldus", 
               "Positiivseid vastandamisi", "Negatiivseid vastandamisi", "Vastandamisi kokku", "Hetketase"]
        sheet.append(headers)

except FileNotFoundError:
    wb = op.Workbook()
    sheet = wb.active  # Kasutan aktiivlehte
    sheet.title = sheet_name 
    headers = ["Reegli ärivõti", "Dimensioon", "Probleemi liik", "Reegli kirjeldus", 
               "Positiivseid vastandamisi", "Negatiivseid vastandamisi", "Vastandamisi kokku", "Hetketase"]
    sheet.append(headers) # Lisan päised

    wb.save(excel_file)

#----------------------------------------------------------------------------------------------------------
  
# Reegel

rule = 'exists'

reegli_kirjeldus_str = 'Kohustuslik atribut peab olema väärtustatud.'

dimensioon_str = "Täelikkus"

probleemi_liik_str = "Puuduv väärtus"

column_list = ['Programm',	'Seiretöö_nimetus', 'Vastutav_partner', 'Vastutav_isik', 'Seirekoha_KKR', 'Seirekoha_nimi', 'Seirekoha_staatus', 'Näitaja_nimetus', 'Proovi_/_vaatluse_kood', 'Proovivõtja_/_vaatlejad',  'Väärtuse_staatus', 'Väärtuse_täpsustus']


with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:

        table[f"{rule}_{column}"] = table[column].apply(exists_check)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'Kontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )
        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        # Kirjutan tulemused tabeli
        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

# Sõnastik hulkaga

valid_names_dict = {
    'Programm': {'C1 Võra seisundi ja kahjustuste hindamine I aste, CC Võra seisundi ja kahjustuste hindamine II aste, DP Sademete seire ja saastekoormus sademetest, FO Okaste ja lehtede seire, LF Varise seire, Metsaseire, SO Metsamulla seire, SS Mullavee seire'},
    'Seiretöö_nimetus': {'Metsaseire 2022 a'},
    'Vastutav_partner': {'Keskkonnaagentuur'},
    'Vastutav_isik': {'Vladislav Apuhtin'},
    'Seirekoha_KKR': {'SJA8621000', 'SJA9826000', 'SJA8221000', 'SJA0933000', 'SJA3587000', 'SJA4510000', 'SJA8645000', 'SJA3859000', 'SJA8505000', 'SJA2903000', 'SJA3702000', 'SJA4933000', 'SJA4147000', 'SJA2883000', 'SJA3745000', 'SJA1342000', 'SJA0988000', 'SJA8924000', 'SJA4098000', 'SJA6403000', 'SJA2466000', 'SJA3307000', 'SJA8794000', 'SJA1659000', 'SJA1901000', 'SJA1932000', 'SJA7506000', 'SJA6096000', 'SJA5462000', 'SJA3617000', 'SJA3306000', 'SJA4215000', 'SJA4228000', 'SJA9530000', 'SJA9174000', 'SJA0129000', 'SJA5508000', 'SJA0027000', 'SJA6249000', 'SJA4096000', 'SJA3024000', 'SJA2324000', 'SJA1654000', 'SJA1837000', 'SJA7022000', 'SJA9926000', 'SJA7136000', 'SJA3623000', 'SJA8167000', 'SJA8279000', 'SJA6495000', 'SJA6570000', 'SJA2809000', 'SJA5182000', 'SJA3752000', 'SJA5944000', 'SJA4553000', 'SJA3403000', 'SJA2260000', 'SJA3611000', 'SJA8232000', 'SJA0397000', 'SJA0242000', 'SJA6027000', 'SJA6265000', 'SJA6033000', 'SJA0892000', 'SJA7920000', 'SJA0925000', 'SJA6874000', 'SJA6238000', 'SJA4772000', 'SJA0483000', 'SJA3302000', 'SJA3547000', 'SJA7182000', 'SJA0040000', 'SJA4966000', 'SJA8412000', 'SJA5689000', 'SJA2390000', 'SJA8145000', 'SJA6111000', 'SJA7863000', 'SJA1825000', 'SJA7574000', 'SJA2066000', 'SJA0141000', 'SJA1531000', 'SJA9412000', 'SJA2722000', 'SJA0808000', 'SJA8426000', 'SJA8385000', 'SJA5547000', 'SJA5144000', 'SJA2847000', 'SJA9611000', 'SJA1609000', 'SJA2755000'},
    'Seirekoha_nimi': {'147', '168', 'II-9 Tõravere (Tartumaa metskond kvartal PE040 eraldis 17)', '4', '70', '67', '28', '104', '183', '194', '59', '1', '40', '165', '151', '25', '2', '115', '46', '170', '187', '140', '174', '114', '108', '192', '120', '162', '45', '41', '127', '103', '98', '7', '131', 'II-2 Vihula (Lääne-Virumaa metskond kvartal VU003 eraldis 12)', '47', '44', '27', '62', '135', '157', '191', '190', '81', '125', '133', '3', '185', '86', '23', '31', '193', '88', '21', '101', '102', '121', '152', '158', '195', '65', '30', '85', '186', '169', '117', 'II-3 Pikasilla (Valgamaa metskond kvartal AA125 eraldis 3)', '153', '124', '107', '43', 'II-8 Karepa (Lääne-Virumaa metskond kvartal QN042 eraldis 5)', '189', '188', '172', '80', '76', '26', '84', 'II-7 Karula (Võrumaa metskond kvartal AS159 eraldised 6, 7)', '63', '167', '64', '13', '38', '175', '141', '109', '29', '129', '20', '35', '57', '154', '5', '42', 'II-1 Sagadi (Lääne-Virumaa metskond kvartal SG106 eraldis 4)', '171', '130'},
    'Seirekoha_staatus': {'Arhiveeritud', 'Kehtiv'}, 
    'Näitaja_nimetus': {'Üldfosfor', 'Võrdluspuu', 'Elavhõbe (kõik puuliigid)', 'Okastiku vanuseklassid', 'Kaalium', 'Ladva seisund', 'Mangaan', 'Üldfosfor (kõik puuliigid)', 'Raud', 'Asendusmagneesium', 'Puu seisund', 'Üldlämmastik', 'Kahjustuse avaldumine', 'Vanuse määramise meetod', 'Okka/lehekadu kogu võra ulatuses', 'Puu kõrgus', 'Okka/lehekadu võra ülemises 1/3 osas', 'Kahjustuse põhjuse nimi', 'Kaadmium (kõik puuliigid)', 'Üldorgaaniline süsinik', 'Vask', 'Tsink', 'Orgaanilise kihi kuivkaal', 'Asendusnaatrium', 'Kaltsium (kõik puuliigid)', 'Üldorgaaniline süsinik (kõik puuliigid)', 'Väävel', 'Kaltsium', 'Nitraatlämmastik (NO3N)', 'Magneesium (kõik puuliigid)', 'Vaadeldav võra osa', 'Raud (kõik puuliigid)', 'Lahustunud orgaaniline süsinik', 'Vanimad okkad', 'Asendushappesus', 'Kahjustuse sümptom', 'Lisavõrsete hulk', 'Boor', 'Üldlämmastik (kõik puuliigid)', 'Kahjustuse vanus', 'Puu vanuseklass', 'Sademete hulk', 'Sulfaatväävel (SO4S)', 'Käbikandvus', 'Kaalium (kõik puuliigid)', 'Tsink (kõik puuliigid)', 'Fosfor', 'Asenduskaalium', 'Kahjustuse koht võras', 'pH (H2O)', 'Kaadmium', 'Puistu I rinde keskmine vanuseklass', 'Magneesium', 'Plii (kõik puuliigid)', 'Kahjustuse ulatus', 'Mullavett proovis', 'Karbonaadid', 'Võra varjutatus', 'Kroom (kõik puuliigid)', 'pH (CaCl2)', 'Elektrijuhtivus', 'Asendusalumiinium', 'Boor (kõik puuliigid)', 'Kahjustuse põhjus', 'Kroom', 'pH', 'Ammooniumlämmastik (NH4N)', 'Naatrium', '1000 okka kuivkaal', 'Alumiinium', 'Mangaan (kõik puuliigid)', 'Puu rinnasdiameeter', 'Asendusraud', 'Leelisus', 'Puu kahjustatud osa', 'Väävel (kõik puuliigid)', 'Lõimis (tekstuurne klass WRB järgi)', 'Õitsemine', 'Nikkel (kõik puuliigid)', 'Elavhõbe', 'Puu kasvuklass', 'Varise kuivkaal m² kohta (kõik puuliigid)', 'Varise kuivkaal m² kohta', 'Nikkel', 'Aktiivne happesus', 'Võra nähtavus', 'Asendusmangaan', 'Mullaniiskus', 'Plii', 'Vask (kõik puuliigid)', 'Puu vanus', 'Kloriid', 'Asenduskaltsium'}, 
    'Vaatlusgrupp': {np.nan, 'Visuaalne puude võra seisundi hindamine'}, 
    'Väärtuse_staatus': {'Kehtiv'}, 
    'Erimärk': {np.nan, '<'}, 
    'Mõõdetud_väärtuse_ühik': {'g/100 g KA', 'cmol+/kg KA', 'ml', np.nan, 'mg/l', 'ng/g KA', 'mg/kg KA', '% KA', 'a', 'mgS/l', 'kg/m² KA', 'µS/cm', 'kg/m²', 'mgC/l', 'μg/g KA', 'µg/l', 'g/kg KA', 'µekv/l', 'g', 'mgN/l', 'mg/g KA', 'm', 'mm', 'cm'}, 
    'Väärtuse_ühik': {'g/100 g KA', 'cmol+/kg KA', 'ml', np.nan, 'mg/l', 'ng/g KA', 'mg/kg KA', '% KA', 'a', 'mgS/l', 'kg/m² KA', 'µS/cm', 'kg/m²', 'mgC/l', 'μg/g KA', 'µg/l', 'g/kg KA', 'µekv/l', 'g', 'mgN/l', 'mg/g KA', 'm', 'mm', 'cm'},
    'Analüüsimeetodi_standard': {'ISO 11466', 'EVS-EN ISO 13395', 'EVS-EN 13137', 'EVS-EN ISO 14911', 'ISO 11465', 'EVS-EN ISO 11732', 'EVS-EN ISO 17294-2', np.nan, 'EVS-EN 27888', 'ISO 11261', 'ISO 10390', 'EVS-EN ISO 6878, sec 7', 'ISO 10694', 'EVS-EN ISO 9963-1', 'ISO 10523', 'ISO 11277', 'ISO 29441', 'ISO 15681-2', 'EVS-EN 1484', 'ISO 11260', 'EVS-EN ISO 10304-1', 'EVS-EN ISO 11885'}, 
    'Analüüsimeetodi_nimetus': {'EVS-EN 1484 (IR)', 'STJnr.M/U94A (ICP-MS)', 'EVS-EN ISO 6878, sec 7 (UV-VIS)', np.nan, 'ISO 11466 (ICP-OES)', 'EVS-EN ISO 14911 (IC)', 'EVS-EN 27888 (EC)', 'ISO 10523 (EC)', 'EVS-EN ISO 9963-1 (TITR)', 'EVS-EN ISO 17294-2 (ICP-MS)', 'ISO 10694 (IR)', 'EVS-EN ISO 11732 (UV-VIS)', 'ISO 10390 (EC)', 'Mulla lõimise määramine - ISO 11277', 'ISO 11260 & ISO 14254 (TITR)', 'ISO 11465 (GR)', 'ISO 15681-2 (UV-VIS)', 'STJnr.M/U84-2A (AFS)', 'Kjeldahl-i meetod ISO 11261 (TITR)', 'EVS-EN ISO 10304-1 (IC)', 'EVS-EN ISO 11885 (ICP-OES)', 'EVS-EN ISO 13395 (UV-VIS)', 'ISO 11260 & ISO 14254 (ICP-OES)', 'ISO 29441 (UV-VIS)', 'STJnr.M/U91 (ICP-OES)', 'EVS-EN 13137 (IR)'}, 
    'Analüüsimeetodi_tööjuhendi_nr': {'STJnr.M/U91',np.nan, 'STJnr.M/U84-2A', 'STJnr.M/U94A'},
    'Analüüsimeetodi_allikas': {np.nan, 'Vee,muda ja settee metallisisalduste määramine induktiivsisestunud plasma aatomemissioonspektromeetriga.'}, 
    'Väärtuse_täpsustus': {' - '}}

rule = 'patern'

reegli_kirjeldus_str = 'Andmeelement peab vastama etteantud loendile.'

dimensioon_str = "Õigsus"

probleemi_liik_str = "Väärtusvahemiku rikkumine"

column_list = ['Programm', 'Seiretöö_nimetus', 'Vastutav_partner', 'Vastutav_isik', 'Seirekoha_KKR', 
                'Seirekoha_nimi', 'Seirekoha_staatus', 'Näitaja_nimetus', 'Vaatlusgrupp', 'Väärtuse_staatus',
                 'Erimärk', 'Mõõdetud_väärtuse_ühik', 'Väärtuse_ühik', 'Analüüsimeetodi_standard',
       'Analüüsimeetodi_nimetus', 'Analüüsimeetodi_tööjuhendi_nr', 'Analüüsimeetodi_allikas', 'Väärtuse_täpsustus']

with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table[column].apply(lambda x: valid_names_check(x, valid_names_dict[column]))

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count       

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )
        #print(result_str)

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

rule = 'formaat'

reegli_kirjeldus_str = 'Seirekoha kood koosneb kolmest suurest tähest ja seitsmest numbrist'

dimensioon_str = "Reeglipärasus"

probleemi_liik_str = "Andmemustritest kõrvalekalded"

column_list = ['Seirekoha_KKR']


with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table[column].apply(format_seirekoha_kkr_check)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

Näitaja_lühend_dict= {'1000 okka kuivkaal': {np.nan},
             'Aktiivne happesus': {'vaba H+'},
             'Alumiinium': {'Al'},
             'Ammooniumlämmastik (NH4N)': {'NH4N'},
             'Asendusalumiinium': {'Al-ex'},
             'Asendushappesus': {'Al+H'},
             'Asenduskaalium': {'K-ex'},
             'Asenduskaltsium': {'Ca-ex'},
             'Asendusmagneesium': {'Mg-ex'},
             'Asendusmangaan': {'Mn-ex'},
             'Asendusnaatrium': {'Na-ex'},
             'Asendusraud': {'Fe-ex'},
             'Boor': {'B'},
             'Boor (kõik puuliigid)': {'B'},
             'Elavhõbe': {'Hg'},
             'Elavhõbe (kõik puuliigid)': {'Hg'},
             'Elektrijuhtivus': {np.nan},
             'Fosfor': {'P'},
             'Kaadmium': {'Cd'},
             'Kaadmium (kõik puuliigid)': {'Cd'},
             'Kaalium': {'K'},
             'Kaalium (kõik puuliigid)': {'K'},
             'Kahjustuse avaldumine': {np.nan},
             'Kahjustuse koht võras': {np.nan},
             'Kahjustuse põhjus': {np.nan},
             'Kahjustuse põhjuse nimi': {np.nan},
             'Kahjustuse sümptom': {np.nan},
             'Kahjustuse ulatus': {np.nan},
             'Kahjustuse vanus': {np.nan},
             'Kaltsium': {'Ca'},
             'Kaltsium (kõik puuliigid)': {'Ca'},
             'Karbonaadid': {np.nan},
             'Kloriid': {'Cl'},
             'Kroom': {'Cr'},
             'Kroom (kõik puuliigid)': {'Cr'},
             'Käbikandvus': {np.nan},
             'Ladva seisund': {np.nan},
             'Lahustunud orgaaniline süsinik': {'DOC'},
             'Leelisus': {np.nan, 'Alka'},
             'Lisavõrsete hulk': {np.nan},
             'Lõimis (tekstuurne klass WRB järgi)': {np.nan},
             'Magneesium': {'Mg'},
             'Magneesium (kõik puuliigid)': {'Mg'},
             'Mangaan': {'Mn'},
             'Mangaan (kõik puuliigid)': {'Mn'},
             'Mullaniiskus': {np.nan},
             'Mullavett proovis': {np.nan},
             'Naatrium': {'Na'},
             'Nikkel': {'Ni'},
             'Nikkel (kõik puuliigid)': {'Ni'},
             'Nitraatlämmastik (NO3N)': {'NO3N'},
             'Okastiku vanuseklassid': {np.nan},
             'Okka/lehekadu kogu võra ulatuses': {np.nan},
             'Okka/lehekadu võra ülemises 1/3 osas': {np.nan},
             'Orgaanilise kihi kuivkaal': {np.nan},
             'Plii': {'Pb'},
             'Plii (kõik puuliigid)': {'Pb'},
             'Puistu I rinde keskmine vanuseklass': {np.nan},
             'Puu kahjustatud osa': {np.nan},
             'Puu kasvuklass': {np.nan},
             'Puu kõrgus': {np.nan},
             'Puu rinnasdiameeter': {np.nan},
             'Puu seisund': {np.nan},
             'Puu vanus': {np.nan},
             'Puu vanuseklass': {np.nan},
             'Raud': {'Fe'},
             'Raud (kõik puuliigid)': {'Fe'},
             'Sademete hulk': {np.nan},
             'Sulfaatväävel (SO4S)': {'SO4S'},
             'Tsink': {'Zn'},
             'Tsink (kõik puuliigid)': {'Zn'},
             'Vaadeldav võra osa': {np.nan},
             'Vanimad okkad': {np.nan},
             'Vanuse määramise meetod': {np.nan},
             'Varise kuivkaal m² kohta': {np.nan},
             'Varise kuivkaal m² kohta (kõik puuliigid)': {np.nan},
             'Vask': {'Cu'},
             'Vask (kõik puuliigid)': {'Cu'},
             'Väävel': {'S'},
             'Väävel (kõik puuliigid)': {'S'},
             'Võra nähtavus': {np.nan},
             'Võra varjutatus': {np.nan},
             'Võrdluspuu': {np.nan},
             'pH': {'pH'},
             'pH (CaCl2)': {'pH (CaCl2)'},
             'pH (H2O)': {'pH (H2O)'},
             'Õitsemine': {np.nan},
             'Üldfosfor': {'Püld'},
             'Üldfosfor (kõik puuliigid)': {'Püld'},
             'Üldlämmastik': {'Nüld'},
             'Üldlämmastik (kõik puuliigid)': {'Nüld'},
             'Üldorgaaniline süsinik': {'TOC'},
             'Üldorgaaniline süsinik (kõik puuliigid)': {'TOC'}}

Näitaja_grupp_dict = {'Füüsikalis-keemilised näitajad elustikus': {'1000 okka kuivkaal',
                                                          'Kaalium',
                                                          'Kaalium (kõik '
                                                          'puuliigid)',
                                                          'Kaltsium',
                                                          'Kaltsium (kõik '
                                                          'puuliigid)',
                                                          'Magneesium',
                                                          'Magneesium (kõik '
                                                          'puuliigid)',
                                                          'Mangaan',
                                                          'Mangaan (kõik '
                                                          'puuliigid)',
                                                          'Raud',
                                                          'Raud (kõik '
                                                          'puuliigid)',
                                                          'Varise kuivkaal m² '
                                                          'kohta',
                                                          'Varise kuivkaal m² '
                                                          'kohta (kõik '
                                                          'puuliigid)',
                                                          'Väävel',
                                                          'Väävel (kõik '
                                                          'puuliigid)',
                                                          'Üldfosfor',
                                                          'Üldfosfor (kõik '
                                                          'puuliigid)',
                                                          'Üldlämmastik',
                                                          'Üldlämmastik (kõik '
                                                          'puuliigid)',
                                                          'Üldorgaaniline '
                                                          'süsinik',
                                                          'Üldorgaaniline '
                                                          'süsinik (kõik '
                                                          'puuliigid)'},
             'Füüsikalis-keemilised näitajad setetes, mullas, pinnases': {'Aktiivne '
                                                                          'happesus',
                                                                          'Asendusalumiinium',
                                                                          'Asendushappesus',
                                                                          'Asenduskaalium',
                                                                          'Asenduskaltsium',
                                                                          'Asendusmagneesium',
                                                                          'Asendusmangaan',
                                                                          'Asendusnaatrium',
                                                                          'Asendusraud',
                                                                          'Fosfor',
                                                                          'Kaalium',
                                                                          'Kaltsium',
                                                                          'Karbonaadid',
                                                                          'Magneesium',
                                                                          'Mangaan',
                                                                          'Orgaanilise '
                                                                          'kihi '
                                                                          'kuivkaal',
                                                                          'pH '
                                                                          '(CaCl2)',
                                                                          'pH '
                                                                          '(H2O)',
                                                                          'Üldlämmastik',
                                                                          'Üldorgaaniline '
                                                                          'süsinik'},
             'Füüsikalis-keemilised näitajad vees': {'Alumiinium',
                                                     'Ammooniumlämmastik '
                                                     '(NH4N)',
                                                     'Elektrijuhtivus',
                                                     'Kaalium',
                                                     'Kaltsium',
                                                     'Kloriid',
                                                     'Lahustunud orgaaniline '
                                                     'süsinik',
                                                     'Leelisus',
                                                     'Magneesium',
                                                     'Mangaan',
                                                     'Mullavett proovis',
                                                     'Naatrium',
                                                     'Nitraatlämmastik (NO3N)',
                                                     'Raud',
                                                     'Sademete hulk',
                                                     'Sulfaatväävel (SO4S)',
                                                     'pH',
                                                     'Üldfosfor',
                                                     'Üldlämmastik'},
             'Muld': {'Lõimis (tekstuurne klass WRB järgi)'},
             'Muud': {'Mullaniiskus'},
             'Ohtlikud ained elustikus': {'Boor',
                                          'Boor (kõik puuliigid)',
                                          'Elavhõbe',
                                          'Elavhõbe (kõik puuliigid)',
                                          'Kaadmium',
                                          'Kaadmium (kõik puuliigid)',
                                          'Kroom',
                                          'Kroom (kõik puuliigid)',
                                          'Nikkel',
                                          'Nikkel (kõik puuliigid)',
                                          'Plii',
                                          'Plii (kõik puuliigid)',
                                          'Tsink',
                                          'Tsink (kõik puuliigid)',
                                          'Vask',
                                          'Vask (kõik puuliigid)'},
             'Ohtlikud ained setetes, mullas, pinnases': {'Kaadmium',
                                                          'Plii',
                                                          'Tsink',
                                                          'Vask'},
             'Ohtlikud ained vees': {'Plii', 'Kaadmium', 'Vask', 'Tsink'},
             'Puistu ja puude näitajad': {'Kahjustuse avaldumine',
                                          'Kahjustuse koht võras',
                                          'Kahjustuse põhjus',
                                          'Kahjustuse põhjuse nimi',
                                          'Kahjustuse sümptom',
                                          'Kahjustuse ulatus',
                                          'Kahjustuse vanus',
                                          'Käbikandvus',
                                          'Ladva seisund',
                                          'Lisavõrsete hulk',
                                          'Okastiku vanuseklassid',
                                          'Okka/lehekadu kogu võra ulatuses',
                                          'Okka/lehekadu võra ülemises 1/3 '
                                          'osas',
                                          'Puu kahjustatud osa',
                                          'Puu kasvuklass',
                                          'Puu kõrgus',
                                          'Puu rinnasdiameeter',
                                          'Puu seisund',
                                          'Puu vanus',
                                          'Puu vanuseklass',
                                          'Vaadeldav võra osa',
                                          'Vanimad okkad',
                                          'Vanuse määramise meetod',
                                          'Võra nähtavus',
                                          'Võra varjutatus',
                                          'Võrdluspuu',
                                          'Õitsemine'},
             'Seirekoha kirjeldused': {'Puistu I rinde keskmine vanuseklass'}}

Alamgrupp_dict = {'Metallid': {'Alumiinium',
                          'Boor',
                          'Boor (kõik puuliigid)',
                          'Elavhõbe',
                          'Elavhõbe (kõik puuliigid)',
                          'Kaadmium',
                          'Kaadmium (kõik puuliigid)',
                          'Kroom',
                          'Kroom (kõik puuliigid)',
                          'Mangaan',
                          'Mangaan (kõik puuliigid)',
                          'Nikkel',
                          'Nikkel (kõik puuliigid)',
                          'Plii',
                          'Plii (kõik puuliigid)',
                          'Tsink',
                          'Tsink (kõik puuliigid)',
                          'Vask',
                          'Vask (kõik puuliigid)'}}

Liik_takson_est_dict ={'arukask': {'Betula pendula'},
             'hall lepp': {'Alnus incana'},
             'harilik haab': {'Populus tremula'},
             'harilik kuusk': {'Picea abies'},
             'harilik mänd': {'Pinus sylvestris'},
             'harilik saar': {'Fraxinus excelsior'},
             'harilik tamm': {'Quercus robur'},
             'sanglepp': {'Alnus glutinosa'},
             'sookask': {'Betula pubescens'}}
Liik_takson_lad_dict= {'Alnus glutinosa': {'Alnus glutinosa'},
             'Alnus incana': {'Alnus incana'},
             'Betula pendula': {'Betula pendula'},
             'Betula pubescens': {'Betula pubescens'},
             'Fraxinus excelsior': {'Fraxinus excelsior'},
             'Picea abies': {'Picea abies'},
             'Pinus sylvestris': {'Pinus sylvestris'},
             'Populus tremula': {'Populus tremula'},
             'Quercus robur': {'Quercus robur'}}

Prooviprotokolli_vaatluslehe_nr_dict = {11.0: {'P0000790227'},
             12.0: {'P0000790238'},
             13.0: {'P0000790188'},
             14.0: {'P0000790234'},
             15.0: {'P0000790167'},
             16.0: {'P0000790210'},
             17.0: {'P0000790163'},
             18.0: {'P0000790119'},
             19.0: {'P0000790129'},
             20.0: {'P0000790250'},
             21.0: {'P0000790222'},
             22.0: {'P0000790218'},
             1830.0: {'P0000790318'},
             1832.0: {'P0000790310'},
             2208.0: {'P0000790186'},
             2209.0: {'P0000790213'},
             2230.0: {'P0000790139'},
             2231.0: {'P0000790251'},
             20.0: {'P0000790250'},
             21.0: {'P0000790222'},
             22.0: {'P0000790218'},
             1830.0: {'P0000790318'},
             1832.0: {'P0000790310'},
             2208.0: {'P0000790186'},
             2209.0: {'P0000790213'},
             2230.0: {'P0000790139'},
             2231.0: {'P0000790251'},
             22.0: {'P0000790218'},
             1830.0: {'P0000790318'},
             1832.0: {'P0000790310'},
             2208.0: {'P0000790186'},
             2209.0: {'P0000790213'},
             2230.0: {'P0000790139'},
             2231.0: {'P0000790251'},
             1832.0: {'P0000790310'},
             2208.0: {'P0000790186'},
             2209.0: {'P0000790213'},
             2230.0: {'P0000790139'},
             2231.0: {'P0000790251'},
             2209.0: {'P0000790213'},
             2230.0: {'P0000790139'},
             2231.0: {'P0000790251'},
             2231.0: {'P0000790251'},
             2232.0: {'P0000790130'},
             2232.0: {'P0000790130'},
             2233.0: {'P0000790193'},
             2234.0: {'P0000790225'},
             2234.0: {'P0000790225'},
             2235.0: {'P0000790204'},
             2235.0: {'P0000790204'},
             2236.0: {'P0000790191'},
             2236.0: {'P0000790191'},
             2237.0: {'P0000790125'},
             2238.0: {'P0000790244'},
             2237.0: {'P0000790125'},
             2238.0: {'P0000790244'},
             2238.0: {'P0000790244'},
             2239.0: {'P0000790118'},
             9611.0: {'P0000790294'}}

Seirekoha_KKR_dict =  {'SJA0027000': {'186'},
             'SJA0040000': {'124'},
             'SJA0129000': {'103'},
             'SJA0141000': {'131'},
             'SJA0242000': {'30'},
             'SJA0397000': {'38'},
             'SJA0483000': {'102'},
             'SJA0808000': {'165'},
             'SJA0892000': {'152'},
             'SJA0925000': {'29'},
             'SJA0933000': {'175'},
             'SJA0988000': {'43'},
             'SJA1342000': {'153'},
             'SJA1531000': {'II-7 Karula (Võrumaa metskond kvartal AS159 '
                            'eraldised 6, 7)'},
             'SJA1609000': {'28'},
             'SJA1654000': {'85'},
             'SJA1659000': {'174'},
             'SJA1825000': {'21'},
             'SJA1837000': {'101'},
             'SJA1901000': {'57'},
             'SJA1932000': {'II-9 Tõravere (Tartumaa metskond kvartal PE040 '
                            'eraldis 17)'},
             'SJA2066000': {'170'},
             'SJA2260000': {'168'},
             'SJA2324000': {'147'},
             'SJA2390000': {'127'},
             'SJA2466000': {'141'},
             'SJA2722000': {'157'},
             'SJA2755000': {'45'},
             'SJA2809000': {'195'},
             'SJA2847000': {'151'},
             'SJA2883000': {'189'},
             'SJA2903000': {'194'},
             'SJA3024000': {'191'},
             'SJA3302000': {'35'},
             'SJA3306000': {'190'},
             'SJA3307000': {'188'},
             'SJA3403000': {'108'},
             'SJA3547000': {'64'},
             'SJA3587000': {'192'},
             'SJA3611000': {'133'},
             'SJA3617000': {'135'},
             'SJA3623000': {'185'},
             'SJA3702000': {'125'},
             'SJA3745000': {'115'},
             'SJA3752000': {'129'},
             'SJA3859000': {'88'},
             'SJA4096000': {'86'},
             'SJA4098000': {'4'},
             'SJA4147000': {'158'},
             'SJA4215000': {'109'},
             'SJA4228000': {'187'},
             'SJA4510000': {'81'},
             'SJA4553000': {'107'},
             'SJA4772000': {'167'},
             'SJA4933000': {'II-3 Pikasilla (Valgamaa metskond kvartal AA125 '
                            'eraldis 3)'},
             'SJA4966000': {'5'},
             'SJA5144000': {'44'},
             'SJA5182000': {'23'},
             'SJA5462000': {'42'},
             'SJA5508000': {'70'},
             'SJA5547000': {'41'},
             'SJA5689000': {'76'},
             'SJA5944000': {'25'},
             'SJA6027000': {'40'},
             'SJA6033000': {'67'},
             'SJA6096000': {'1'},
             'SJA6111000': {'121'},
             'SJA6238000': {'117'},
             'SJA6249000': {'120'},
             'SJA6265000': {'47'},
             'SJA6403000': {'183'},
             'SJA6495000': {'63'},
             'SJA6570000': {'3'},
             'SJA6874000': {'140'},
             'SJA7022000': {'31'},
             'SJA7136000': {'193'},
             'SJA7182000': {'65'},
             'SJA7506000': {'II-8 Karepa (Lääne-Virumaa metskond kvartal QN042 '
                            'eraldis 5)'},
             'SJA7574000': {'172'},
             'SJA7863000': {'II-2 Vihula (Lääne-Virumaa metskond kvartal VU003 '
                            'eraldis 12)'},
             'SJA7920000': {'13'},
             'SJA8145000': {'169'},
             'SJA8167000': {'20'},
             'SJA8221000': {'2'},
             'SJA8232000': {'80'},
             'SJA8279000': {'171'},
             'SJA8385000': {'84'},
             'SJA8412000': {'26'},
             'SJA8426000': {'130'},
             'SJA8505000': {'46'},
             'SJA8621000': {'104'},
             'SJA8645000': {'7'},
             'SJA8794000': {'27'},
             'SJA8924000': {'154'},
             'SJA9174000': {'II-1 Sagadi (Lääne-Virumaa metskond kvartal SG106 '
                            'eraldis 4)'},
             'SJA9412000': {'114'},
             'SJA9530000': {'162'},
             'SJA9611000': {'98'},
             'SJA9826000': {'62'},
             'SJA9926000': {'59'}}

Vaatleja_EELIS_ID_dict = {
            'Enn Kaljula': {'-122331639.0'},
             'Eve Kaur': {np.nan},
             'Kaarel Aruste': {'-2056850093.0'},
             'Merit Ehrpais': {'2092465016.0'},
             'Teadmata': {np.nan},
             'Vladislav Apuhtin': {np.nan, '190547565.0'}}

Proovivõtumeetodi_nimetus_dict =  {np.nan: {np.nan},
             'Keskmistatud proov': {'Mullavee proovide võtmine lüsimeetrite '
                                    'abil',
                                    'Proovide võtmine sademeteveest '
                                    '(avamaavesi)',
                                    'Proovide võtmine sademeteveest (võravesi)',
                                    'Variseproov (muude puuliikide okkad ja '
                                    'lehed)',
                                    'Variseproov (oksad ja muu varis)',
                                    'Variseproov (peapuuliigi okkad)',
                                    'Variseproov (viljad ja seemned)'},
             'Koondproov': {'Proovivõtt kihtide kaupa - metsamullad'}}

Proovi_liik_dict = {np.nan: {np.nan},
             'Keskmistatud proov': {'Varis', 'Mullavesi', 'Sademete vesi'},
             'Koondproov': {'Muld, pinnas'}}

Proovimaatriks_dict = {np.nan: {np.nan},
             'Muld, pinnas': {'Proovivõtt kihtide kaupa - metsamullad'},
             'Mullavesi': {'Mullavee proovide võtmine lüsimeetrite abil'},
             'Sademete vesi': {'Proovide võtmine sademeteveest (avamaavesi)',
                               'Proovide võtmine sademeteveest (võravesi)'},
             'Varis': {'Variseproov (muude puuliikide okkad ja lehed)',
                       'Variseproov (oksad ja muu varis)',
                       'Variseproov (peapuuliigi okkad)',
                       'Variseproov (viljad ja seemned)'}}

Mullaproov_dict = {np.nan: {np.nan},
             'Kihist': {'H01',
                        'H12',
                        'H24',
                        'H48',
                        'HF',
                        'HFS',
                        'HS',
                        'M05',
                        'M12',
                        'M24',
                        'M48',
                        'M51',
                        'OF',
                        'OH',
                        'OLF'}}

Mullahorisont_dict = {np.nan: {np.nan},
             'H01': {'0.0 - 10.0'},
             'H12': {'10.0 - 20.0'},
             'H24': {'20.0 - 40.0'},
             'H48': {'40.0 - 80.0'},
             'HF': {'-15.0 - -11.0', '-16.0 - -13.0'},
             'HFS': {'-13.0 - -6.0', '-11.0 - -6.0'},
             'HS': {'-6.0 - 0.0'},
             'M05': {'0.0 - 5.0'},
             'M12': {'10.0 - 20.0'},
             'M24': {'20.0 - 40.0'},
             'M48': {'40.0 - 80.0'},
             'M51': {'5.0 - 10.0'},
             'OF': {'-1.0 - 0.0',
                    '-12.0 - -3.0',
                    '-17.0 - -16.0',
                    '-2.0 - 0.0',
                    '-3.0 - 0.0',
                    '-4.0 - 0.0',
                    '-6.0 - -1.0',
                    '-6.0 - -3.0',
                    '-8.0 - -1.0',
                    '-8.0 - -5.0'},
             'OH': {'-1.0 - 0.0', '-3.0 - 0.0', '-5.0 - 0.0'},
             'OLF': {'-1.0 - 0.0'}}

Mullaproovi_sügavus_dict = {np.nan: {np.nan},
             '-1.0 - 0.0': {'Kihist'},
             '-11.0 - -6.0': {'Kihist'},
             '-12.0 - -3.0': {'Kihist'},
             '-13.0 - -6.0': {'Kihist'},
             '-15.0 - -11.0': {'Kihist'},
             '-16.0 - -13.0': {'Kihist'},
             '-17.0 - -16.0': {'Kihist'},
             '-2.0 - 0.0': {'Kihist'},
             '-3.0 - 0.0': {'Kihist'},
             '-4.0 - 0.0': {'Kihist'},
             '-5.0 - 0.0': {'Kihist'},
             '-6.0 - -1.0': {'Kihist'},
             '-6.0 - -3.0': {'Kihist'},
             '-6.0 - 0.0': {'Kihist'},
             '-8.0 - -1.0': {'Kihist'},
             '-8.0 - -5.0': {'Kihist'},
             '0.0 - 10.0': {'Kihist'},
             '0.0 - 5.0': {'Kihist'},
             '10.0 - 20.0': {'Kihist'},
             '20.0 - 40.0': {'Kihist'},
             '40.0 - 80.0': {'Kihist'},
             '5.0 - 10.0': {'Kihist'}}

correlation_dict = {
    'Näitaja_lühend': {'key': 'Näitaja_nimetus', 'value': 'Näitaja_lühend', 'dicti': Näitaja_lühend_dict, 'corr1': 'Näitaja_lühend', 'corr2': 'Näitaja_nimetus'},
    'Näitaja_grupp': {'key': 'Näitaja_grupp', 'value': 'Näitaja_nimetus', 'dicti': Näitaja_grupp_dict, 'corr1': 'Näitaja_grupp', 'corr2': 'Näitaja_nimetus'},
    'Alamgrupp': {'key': 'Alamgrupp', 'value': 'Näitaja_nimetus', 'dicti': Alamgrupp_dict, 'corr1': 'Alamgrupp', 'corr2': 'Näitaja_nimetus'},
    'Liik/takson_(est)': {'key': 'Liik/takson_(est)', 'value': 'Liik/takson_(lad)', 'dicti': Liik_takson_est_dict, 'corr1': 'Liik/takson_(est)', 'corr2': 'Liik/takson_(lad)'},
    'Liik/takson_(lad)': {'key': 'Liik/takson_(lad)', 'value': 'Sisestatud_liik_(lad)', 'dicti': Liik_takson_lad_dict, 'corr1': 'Liik/takson_(lad)', 'corr2': 'Sisestatud_liik_(lad)'},
    'Sisestatud_liik_(lad)': {'key': 'Liik/takson_(lad)', 'value': 'Sisestatud_liik_(lad)', 'dicti': Liik_takson_lad_dict, 'corr1': 'Sisestatud_liik_(lad)', 'corr2': 'Liik/takson_(lad)'},
    'Prooviprotokolli_/_vaatluslehe_nr': {'key': 'Prooviprotokolli_/_vaatluslehe_nr', 'value': 'Proovi_/_vaatluse_kood', 'dicti': Prooviprotokolli_vaatluslehe_nr_dict, 'corr1': 'Prooviprotokolli_/_vaatluslehe_nr', 'corr2': 'Proovi_/_vaatluse_kood'},
    'Seirekoha_KKR': {'key': 'Seirekoha_KKR', 'value': 'Seirekoha_nimi', 'dicti': Seirekoha_KKR_dict, 'corr1': 'Seirekoha_KKR', 'corr2': 'Seirekoha_nimi'},
    'Vaatleja_EELIS_ID': {'key': 'Proovivõtja_/_vaatlejad', 'value': 'Vaatleja_EELIS_ID', 'dicti': Vaatleja_EELIS_ID_dict, 'corr1': 'Vaatleja_EELIS_ID', 'corr2': 'Proovivõtja_/_vaatlejad'},
    'Proovivõtumeetodi_nimetus': {'key': 'Proovi_liik', 'value': 'Proovivõtumeetodi_nimetus', 'dicti': Proovivõtumeetodi_nimetus_dict, 'corr1': 'Proovivõtumeetodi_nimetus', 'corr2': 'Proovi_liik'},
    'Proovi_liik': {'key': 'Proovi_liik', 'value': 'Proovimaatriks', 'dicti': Proovi_liik_dict, 'corr1': 'Proovi_liik', 'corr2': 'Proovimaatriks'},
    'Proovimaatriks': {'key': 'Proovimaatriks', 'value': 'Proovivõtumeetodi_nimetus', 'dicti': Proovimaatriks_dict, 'corr1': 'Proovimaatriks', 'corr2': 'Proovivõtumeetodi_nimetus'},
    'Mullaproov': {'key': 'Mullaproov', 'value': 'Mullahorisont', 'dicti': Mullaproov_dict, 'corr1': 'Mullaproov', 'corr2': 'Mullahorisont'},
    'Mullahorisont': {'key': 'Mullahorisont', 'value': 'Mullaproovi_sügavus', 'dicti': Mullahorisont_dict, 'corr1': 'Mullahorisont', 'corr2': 'Mullaproovi_sügavus'},
    'Mullaproovi_sügavus': {'key': 'Mullaproovi_sügavus', 'value': 'Mullaproov', 'dicti': Mullaproovi_sügavus_dict, 'corr1': 'Mullaproovi_sügavus', 'corr2': 'Mullaproov'}

}

rule = 'correlation'

#reegli_kirjeldus_str = 'Esimene andmeelement peab olema kooskõlas teise elemendiga.

dimensioon_str = "Reeglipärasus"

probleemi_liik_str = "Funktsionaalse sõltuvuse rikkumine"

column_list = ['Näitaja_lühend', 'Näitaja_grupp', 'Alamgrupp', 'Liik/takson_(est)', 'Liik/takson_(lad)',
                'Sisestatud_liik_(lad)', 'Prooviprotokolli_/_vaatluslehe_nr', 'Seirekoha_KKR',
                  'Vaatleja_EELIS_ID', 'Proovivõtumeetodi_nimetus', 'Proovimaatriks', 'Mullaproov',
                   'Mullahorisont', 'Mullaproovi_sügavus' ]


with open(output_file, 'a', encoding='utf-8') as file:
    
    for column in column_list:
        
        # Siin rakendatakse apply() kogu DataFrame'ile
        # axis=1 määrab, et apply() peaks töötlema iga ridu tervikuna),
        #  et lambda pääseks juurde erinevate veergude väärtustele.
        table[f"{rule}_{column}"] = table.apply(lambda row: correlation_check(row[correlation_dict[column]['key']], row[correlation_dict[column]['value']], correlation_dict[column]['dicti']), axis=1)

        passed_count = table[f"{rule}_{column}"].sum()
        failed_count = total_count - passed_count
        excel_passed_percentage = passed_count / total_count
        passed_percentage = (passed_count / total_count) * 100

        file.write('-------------------------------------------\n') 
        file.write(f"{correlation_dict[column]['corr1']} peab olema vastavuses väljaga {correlation_dict[column]['corr2']}.")

        file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"
        f'\nKontrolli läbinud ridade arv: {passed_count}\n'
        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'
        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )
        #print(result_str)

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            f"{correlation_dict[column]['corr1']} peab vastama õigele {correlation_dict[column]['corr2']} väärtusele.",   # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

rule = 'condition'

reegli_kirjeldus_str = 'x L-EST97 peab jääma vahemikku 6383851 ja 6606656'

dimensioon_str = "Õigsus"

probleemi_liik_str = "Väärtusvahemiku rikkumine"

column_list = ['Seirekoha_x_L-EST97', 'Sisestatud_x_L-EST97']

with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table[column].apply(condition_seirekoha_x_check)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

rule = 'condition'

reegli_kirjeldus_str = 'y L-EST97 peab jääma vahemikku 387377 ja 733508'

dimensioon_str = "Õigsus"

probleemi_liik_str = "Väärtusvahemiku rikkumine"

column_list = ['Seirekoha_y_L-EST97', 'Sisestatud_y_L-EST97']


with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table[column].apply(condition_seirekoha_y_check)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

Mõõdetud_arvväärtus_dict = {
    '1000 okka kuivkaal': {
        'g': {'max': 3.97, 'min': 0}
    },
    'Aktiivne happesus': {
        'cmol+/kg KA': {'max': 15.0, 'min': 0}
    },
    'Alumiinium': {
        'mg/l': {'max': 2.7, 'min': 0}
    },
    'Ammooniumlämmastik (NH4N)': {
        'mgN/l': {'max': 4.8, 'min': 0}
    },
    'Asendusalumiinium': {
        'cmol+/kg KA': {'max': 10.0, 'min': 0}
    },
    'Asendushappesus': {
        'cmol+/kg KA': {'max': 22.0, 'min': 0}
    },
    'Asenduskaalium': {
        'cmol+/kg KA': {'max': 2.4, 'min': 0}
    },
    'Asenduskaltsium': {
        'cmol+/kg KA': {'max': 110.0, 'min': 0}
    },
    'Asendusmagneesium': {
        'cmol+/kg KA': {'max': 8.4, 'min': 0}
    },
    'Asendusmangaan': {
        'cmol+/kg KA': {'max': 2.3, 'min': 0}
    },
    'Asendusnaatrium': {
        'cmol+/kg KA': {'max': 1.2, 'min': 0}
    },
    'Asendusraud': {
        'cmol+/kg KA': {'max': 0.62, 'min': 0}
    },
    'Boor': {
        'μg/g KA': {'max': 15.0, 'min': 0}
    },
    'Boor (kõik puuliigid)': {
        'μg/g KA': {'max': 28.0, 'min': 0}
    },
    'Elavhõbe': {
        'ng/g KA': {'max': 27.0, 'min': 0}
    },
    'Elavhõbe (kõik puuliigid)': {
        'ng/g KA': {'max': 93.0, 'min': 0}
    },
    'Elektrijuhtivus': {
        'µS/cm': {'max': 245.0, 'min': 0}
    },
    'Fosfor': {
        'mg/kg KA': {'max': 1600.0, 'min': 0}
    },
    'Kaadmium': {
        'mg/kg KA': {'max': 0.69, 'min': 0},
        'ng/g KA': {'max': 34.0, 'min': 0},
        'µg/l': {'max': 0.26, 'min': 0}
    },
    'Kaadmium (kõik puuliigid)': {
        'ng/g KA': {'max': 180.0, 'min': 0},
    },
    'Kaalium': {
        'mg/g KA': {'max': 4.4, 'min': 0},
        'mg/kg KA': {'max': 1700.0, 'min': 0},
        'mg/l': {'max': 33.0, 'min': 0}
    },
    'Kaalium (kõik puuliigid)': {
        'mg/g KA': {'max': 6.0, 'min': 0}
    },
    'Kaltsium': {
        'mg/g KA': {'max': 13.0, 'min': 0},
        'mg/kg KA': {'max': 41000.0, 'min': 0},
        'mg/l': {'max': 20.0, 'min': 0}
    },
    'Kaltsium (kõik puuliigid)': {
        'mg/g KA': {'max': 16.0, 'min': 0}
    },
    'Karbonaadid': {
        'g/kg KA': {'max': 42.0, 'min': 0}
    },
    'Kloriid': {
        'mg/l': {'max': 29.0, 'min': 0}
    },
    'Kroom': {
        'μg/g KA': {'max': 0.64, 'min': 0}
    },
    'Kroom (kõik puuliigid)': {
        'μg/g KA': {'max': 2.3, 'min': 0}
    },
    'Lahustunud orgaaniline süsinik': {
        'mgC/l': {'max': 95.0, 'min': 0}
    },
    'Leelisus': {
        'µekv/l': {'max': 414.0, 'min': 0}
    },
    'Magneesium': {
        'mg/g KA': {'max': 1.2, 'min': 0},
        'mg/kg KA': {'max': 1800.0, 'min': 0},
        'mg/l': {'max': 7.8, 'min': 0}
    },
    'Magneesium (kõik puuliigid)': {
        'mg/g KA': {'max': 3.1, 'min': 0}
    },
    'Mangaan': {
        'mg/kg KA': {'max': 1400.0, 'min': 0},
        'mg/l': {'max': 0.49, 'min': 0},
        'μg/g KA': {'max': 910.0, 'min': 0}
    },
    'Mangaan (kõik puuliigid)': {
        'μg/g KA': {'max': 940.0, 'min': 0}
    },
    'Mullaniiskus': {
        '% KA': {'max': 11.1, 'min': 0}
    },
    'Mullavett proovis': {
        'ml': {'max': 4866.666667, 'min': 0}
    },
    'Naatrium': {
        'mg/l': {'max': 12.0, 'min': 0}
    },
    'Nikkel': {
        'μg/g KA': {'max': 1.5, 'min': 0}
    },
    'Nikkel (kõik puuliigid)': {
        'μg/g KA': {'max': 2.5, 'min': 0}
    },
    'Nitraatlämmastik (NO3N)': {
        'mgN/l': {'max': 2.5, 'min': 0}
    },
    'Orgaanilise kihi kuivkaal': {
        'kg/m² KA': {'max': 58.62, 'min': 0}
    },
    'Plii': {
        'mg/kg KA': {'max': 77.0, 'min': 0},
        'µg/l': {'max': 11.0, 'min': 0},
        'μg/g KA': {'max': 0.1, 'min': 0}
    },
    'Plii (kõik puuliigid)': {
        'μg/g KA': {'max': 2.4, 'min': 0}
    },
    'Puu kõrgus': {
        'm': {'max': 39.3, 'min': 0}
    },
    'Puu rinnasdiameeter': {
        'cm': {'max': 59.0, 'min': 0}
    },
    'Puu vanus': {
        'a': {'max': 179.0, 'min': 0}
    },
    'Raud': {
        'mg/l': {'max': 0.75, 'min': 0},
        'μg/g KA': {'max': 51.0, 'min': 0}
    },
    'Raud (kõik puuliigid)': {
        'μg/g KA': {'max': 640.0, 'min': 0}
    },
    'Sademete hulk': {
        'mm': {'max': 157.0, 'min': 0}
    },
    'Sulfaatväävel (SO4S)': {
        'mgS/l': {'max': 26.0, 'min': 0}
    },
    'Tsink': {
        'mg/kg KA': {'max': 110.0, 'min': 0},
        'µg/l': {'max': 74.0, 'min': 0},
        'μg/g KA': {'max': 20.0, 'min': 0}
    },
    'Tsink (kõik puuliigid)': {
        'μg/g KA': {'max': 90.0, 'min': 0}
    },
    'Vanimad okkad': {
        'a': {'max': 4.0, 'min': 0}
    },
    'Varise kuivkaal m² kohta': {
        'kg/m²': {'max': 0.714541548, 'min': 0}
    },
    'Varise kuivkaal m² kohta (kõik puuliigid)': {
        'kg/m²': {'max': 0.151646301, 'min': 0}
    },
    'Vask': {
        'mg/kg KA': {'max': 18.0, 'min': 0},
        'µg/l': {'max': 40.0, 'min': 0},
        'μg/g KA': {'max': 2.3, 'min': 0}
    },
    'Vask (kõik puuliigid)': {
        'μg/g KA': {'max': 13.0, 'min': 0}
    },
    'Väävel': {
        'mg/g KA': {'max': 0.87, 'min': 0}
    },
    'Väävel (kõik puuliigid)': {
        'mg/g KA': {'max': 1.3, 'min': 0}
    },
    'Üldfosfor': {
        'mg/g KA': {'max': 1.4, 'min': 0},
        'mg/l': {'max': 3.6, 'min': 0}
    },
    'Üldfosfor (kõik puuliigid)': {
        'mg/g KA': {'max': 1.5, 'min': 0}
    },
    'Üldlämmastik': {
        'g/kg KA': {'max': 33.0, 'min': 0},
        'mg/g KA': {'max': 8.8, 'min': 0},
        'mg/l': {'max': 10.0, 'min': 0}
    },
    'Üldlämmastik (kõik puuliigid)': {
        'mg/g KA': {'max': 12.0, 'min': 0}
    },
    'Üldorgaaniline süsinik': {
        'g/100 g KA': {'max': 48.0, 'min': 0},
        'g/kg KA': {'max': 500.0, 'min': 0},
    },
    'Üldorgaaniline süsinik (kõik puuliigid)': {
        'g/100 g KA': {'max': 50.0, 'min': 0}
    },
    'pH': {
        None: {'max': 14, 'min': 0}
    },
    'pH (CaCl2)': {
        None: {'max': 14, 'min': 0}
    },
    'pH (H2O)': {
        None: {'max': 14, 'min': 0}
    },
    'Okastiku vanuseklassid': {
        None: {'max': 10, 'min': 0}
    }
}



rule = 'condition'

reegli_kirjeldus_str = 'Mõõdetud arväärtus peab olema lubatud piirväärtuste vahemikus.'

dimensioon_str = "Õigsus"

probleemi_liik_str = "Väärtusvahemiku rikkumine"

column_list = ['Mõõdetud_arvväärtus']

with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table.apply(lambda row: min_max_check(row['Näitaja_nimetus'], row['Mõõdetud_arvväärtus'], row['Mõõdetud_väärtuse_ühik'], Mõõdetud_arvväärtus_dict), axis=1)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

Arvväärtus_dict = {
    '1000 okka kuivkaal': {'g': {'max': 3.97, 'min': 0}},
    'Aktiivne happesus': {'cmol+/kg KA': {'max': 15.0, 'min': 0}},
    'Alumiinium': {'mg/l': {'max': 2.7, 'min': 0}},
    'Ammooniumlämmastik (NH4N)': {'mgN/l': {'max': 4.8, 'min': 0}},
    'Asendusalumiinium': {'cmol+/kg KA': {'max': 10.0, 'min': 0}},
    'Asendushappesus': {'cmol+/kg KA': {'max': 22.0, 'min': 0}},
    'Asenduskaalium': {'cmol+/kg KA': {'max': 2.4, 'min': 0}},
    'Asenduskaltsium': {'cmol+/kg KA': {'max': 110.0, 'min': 0}},
    'Asendusmagneesium': {'cmol+/kg KA': {'max': 8.4, 'min': 0}},
    'Asendusmangaan': {'cmol+/kg KA': {'max': 2.3, 'min': 0}},
    'Asendusnaatrium': {'cmol+/kg KA': {'max': 1.2, 'min': 0}},
    'Asendusraud': {'cmol+/kg KA': {'max': 0.62, 'min': 0}},
    'Boor': {'μg/g KA': {'max': 15.0, 'min': 0}},
    'Boor (kõik puuliigid)': {'μg/g KA': {'max': 28.0, 'min': 0}},
    'Elavhõbe': {'mg/kg KA': {'max': 0.027, 'min': 0}},
    'Elavhõbe (kõik puuliigid)': {'mg/kg KA': {'max': 0.093, 'min': 0}},
    'Elektrijuhtivus': {'µS/cm': {'max': 245.0, 'min': 0}},
    'Fosfor': {'mg/kg KA': {'max': 1600.0, 'min': 0}},
    'Kaadmium': {
        'mg/kg KA': {'max': 0.69, 'min': 0},
        'mg/l': {'max': 0.00026, 'min': 0},
        'ng/g KA': {'max': 34.0, 'min': 0}
    },
    'Kaadmium (kõik puuliigid)': {'ng/g KA': {'max': 180.0, 'min': 0}},
    'Kaalium': {
        'mg/g KA': {'max': 4.4, 'min': 0},
        'mg/kg KA': {'max': 1700.0, 'min': 0},
        'mg/l': {'max': 33.0, 'min': 0}
    },
    'Kaalium (kõik puuliigid)': {'mg/g KA': {'max': 6.0, 'min': 0}},
    'Kaltsium': {
        'mg/g KA': {'max': 13.0, 'min': 0},
        'mg/kg KA': {'max': 41000.0, 'min': 0},
        'mg/l': {'max': 20.0, 'min': 0}
    },
    'Kaltsium (kõik puuliigid)': {'mg/g KA': {'max': 16.0, 'min': 0}},
    'Karbonaadid': {'g/kg KA': {'max': 42.0, 'min': 0}},
    'Kloriid': {'mg/l': {'max': 29.0, 'min': 0}},
    'Kroom': {'mg/kg KA': {'max': 0.64, 'min': 0}},
    'Kroom (kõik puuliigid)': {'mg/kg KA': {'max': 2.3, 'min': 0}},
    'Lahustunud orgaaniline süsinik': {
        'mg/l': {'max': 95.0, 'min': 0},
        'mgC/l': {'max': 43.0, 'min': 0}
    },
    'Leelisus': {'µekv/l': {'max': 414.0, 'min': 0}},
    'Magneesium': {
        'mg/g KA': {'max': 1.2, 'min': 0},
        'mg/kg KA': {'max': 1800.0, 'min': 0},
        'mg/l': {'max': 7.8, 'min': 0}
    },
    'Magneesium (kõik puuliigid)': {'mg/g KA': {'max': 3.1, 'min': 0}},
    'Mangaan': {
        'mg/kg KA': {'max': 1400.0, 'min': 0},
        'mg/l': {'max': 0.49, 'min': 0},
        'μg/g KA': {'max': 910.0, 'min': 0}
    },
    'Mangaan (kõik puuliigid)': {'μg/g KA': {'max': 940.0, 'min': 0}},
    'Mullaniiskus': {'% KA': {'max': 11.1, 'min': 0}},
    'Mullavett proovis': {'ml': {'max': 4866.666667, 'min': 0}},
    'Naatrium': {'mg/l': {'max': 12.0, 'min': 0}},
    'Nikkel': {'mg/kg KA': {'max': 1.5, 'min': 0}},
    'Nikkel (kõik puuliigid)': {'mg/kg KA': {'max': 2.5, 'min': 0}},
    'Nitraatlämmastik (NO3N)': {'mgN/l': {'max': 2.5, 'min': 0}},
    'Orgaanilise kihi kuivkaal': {'kg/m² KA': {'max': 58.62, 'min': 0}},
    'Plii': {
        'mg/kg KA': {'max': 77.0, 'min': 0},
        'µg/l': {'max': 11.0, 'min': 0},
        'μg/g KA': {'max': 0.1, 'min': 0}
    },
    'Plii (kõik puuliigid)': {'μg/g KA': {'max': 2.4, 'min': 0}},
    'Puu kõrgus': {'m': {'max': 39.3, 'min': 0}},
    'Puu rinnasdiameeter': {'cm': {'max': 59.0, 'min': 0}},
    'Puu vanus': {'a': {'max': 179.0, 'min': 0}},
    'Raud': {
        'mg/l': {'max': 0.75, 'min': 0},
        'μg/g KA': {'max': 51.0, 'min': 0}
    },
    'Raud (kõik puuliigid)': {'μg/g KA': {'max': 640.0, 'min': 0}},
    'Sademete hulk': {'mm': {'max': 157.0, 'min': 0}},
    'Sulfaatväävel (SO4S)': {'mgS/l': {'max': 26.0, 'min': 0}},
    'Tsink': {
        'mg/kg KA': {'max': 110.0, 'min': 0},
        'µg/l': {'max': 74.0, 'min': 0},
        'μg/g KA': {'max': 20.0, 'min': 0}
    },
    'Tsink (kõik puuliigid)': {'μg/g KA': {'max': 90.0, 'min': 0}},
    'Vanimad okkad': {'a': {'max': 4.0, 'min': 0}},
    'Varise kuivkaal m² kohta': {'kg/m²': {'max': 0.714541548, 'min': 0}},
    'Varise kuivkaal m² kohta (kõik puuliigid)': {'kg/m²': {'max': 0.151646301, 'min': 0}},
    'Vask': {
        'mg/kg KA': {'max': 18.0, 'min': 0},
        'µg/l': {'max': 40.0, 'min': 0},
        'μg/g KA': {'max': 2.3, 'min': 0}
    },
    'Vask (kõik puuliigid)': {'μg/g KA': {'max': 13.0, 'min': 0}},
    'Väävel': {
        'mg/g KA': {'max': 0.87, 'min': 0},
        'mg/g KA (kõik puuliigid)': {'max': 1.3, 'min': 0}
    },
    'Üldfosfor': {
        'mg/g KA': {'max': 1.4, 'min': 0},
        'mg/l': {'max': 3.6, 'min': 0}
    },
    'Üldfosfor (kõik puuliigid)': {'mg/g KA': {'max': 1.5, 'min': 0}},
    'Üldlämmastik': {
        'mg/g KA': {'max': 8.8, 'min': 0},
        'mg/kg KA': {'max': 33000.0, 'min': 0},
        'mg/l': {'max': 10.0, 'min': 0}
    },
    'Üldlämmastik (kõik puuliigid)': {'mg/g KA': {'max': 12.0, 'min': 0}},
    'Üldorgaaniline süsinik': {
        'g/100 g KA': {'max': 48.0, 'min': 0},
        'mg/kg KA': {'max': 500000.0, 'min': 0}
    },
    'Üldorgaaniline süsinik (kõik puuliigid)': {'g/100 g KA': {'max': 50.0, 'min': 0}},
    'pH': {
        np.nan: {'max': 14, 'min': 0}},
    'pH (CaCl2)': {
        np.nan: {'max': 14, 'min': 0}},
    'pH (H2O)': {
        np.nan:{'max': 14, 'min': 0}},
    'Okastiku vanuseklassid': {
        np.nan:{'max': 10, 'min': 0}}
}


rule = 'condition'

reegli_kirjeldus_str = 'Arväärtus peab olema lubatud piirväärtuste vahemikus.'

dimensioon_str = "Õigsus"

probleemi_liik_str = "Väärtusvahemiku rikkumine"

column_list = ['Arvväärtus']

with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table.apply(lambda row: min_max_check(row['Näitaja_nimetus'], row['Arvväärtus'], row['Väärtuse_ühik'], Arvväärtus_dict), axis=1)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#----------------------------------------------------------------------------------------------------------

rule = 'condition'

reegli_kirjeldus_str = 'Kui Arvväärtus väli on tühi, siis peab olema täidetud Väärtus (muu) väli'

dimensioon_str = "Reeglipärasus"

probleemi_liik_str = "Funktsionaalse sõltuvuse rikkumine"

column_list = ['Arvväärtus']

with open(output_file, 'a', encoding='utf-8') as file:
    file.write('-------------------------------------------\n') 
    file.write(reegli_kirjeldus_str)
    file.write(f"\nDimensioon: {dimensioon_str}\nProbleemi liik: {probleemi_liik_str}{'\n\n'}")
    for column in column_list:
        
        table[f"{rule}_{column}"] = table.apply(lambda row: condition_vaartus_check(row['Arvväärtus'], row['Väärtus_(muud)']), axis=1)

        passed_count = table[f"{rule}_{column}"].sum()

        failed_count = total_count - passed_count

        excel_passed_percentage = passed_count / total_count

        passed_percentage = (passed_count / total_count) * 100

        result_str = (
        f"\nReegli ärivõti: {rule}_{column}\n"

        f'\nKontrolli läbinud ridade arv: {passed_count}\n'

        f'Kontrolli pole läbinud ridade arv: {failed_count}\n'

        f'Andmekvaliteedi hetketase: {passed_percentage:.2f}%\n\n'
        )

        # Kirjutan tulemused faili
        file.write(result_str) 

        new_row = [
            f"{rule}_{column}",        # Reegli ärivõti
            dimensioon_str,            # Dimensioon
            probleemi_liik_str,        # Probleemi liik
            reegli_kirjeldus_str,      # Reegli kirjeldus
            passed_count,              # Positiivseid vastandamisi
            failed_count,              # Negatiivseid vastandamisi
            total_count,               # Vastandamisi kokku
            excel_passed_percentage    # Hetketase   
        ]

        sheet.append(new_row)

#==============================================================================================================================

# Määran selle veeru indeksi ja tähe, milles väärtused on protsendivormingus
percent_column_index = 8
column_letter = get_column_letter(percent_column_index)

# Vormingu rakendamine tervele veerule
for row in range(2, sheet.max_row + 1):  # Alustan teisest reast, et päist vahele jätta
    cell = f"{column_letter}{row}"
    sheet[cell].number_format = '0.00%'

wb.save(excel_file)
wb.close()

#==============================================================================================================================
# Õige lehe valimine pealkirja järgi
sheet_name = 'Metsaseire_2022_Tulemused' # lehe nimi

with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    table.to_excel(writer, sheet_name=sheet_name, index=False)

#==============================================================================================================================
# Lõppaeg
end_time = time.time()
print(f"Täitmise aeg: {end_time - start_time} sekundit")